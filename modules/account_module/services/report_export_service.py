from datetime import datetime
from typing import Optional, Tuple
from io import BytesIO
from core.database.connection import db_manager
from modules.account_module.models.entities import AccountMaster, AccountGroup, Ledger
from sqlalchemy import func

class ReportExportService:
    def __init__(self):
        self.session_manager = db_manager.get_session_manager()
    
    def export_trial_balance(self, from_date: Optional[str], to_date: Optional[str], format: str) -> Tuple[bytes, str, str]:
        data = self._get_trial_balance_data(from_date, to_date)
        
        if format == 'pdf':
            return self._generate_trial_balance_pdf(data, from_date, to_date)
        else:
            return self._generate_trial_balance_excel(data, from_date, to_date)
    
    def export_profit_loss(self, from_date: Optional[str], to_date: Optional[str], format: str) -> Tuple[bytes, str, str]:
        data = self._get_profit_loss_data(from_date, to_date)
        
        if format == 'pdf':
            return self._generate_profit_loss_pdf(data, from_date, to_date)
        else:
            return self._generate_profit_loss_excel(data, from_date, to_date)
    
    def export_balance_sheet(self, as_of_date: Optional[str], format: str) -> Tuple[bytes, str, str]:
        data = self._get_balance_sheet_data(as_of_date)
        
        if format == 'pdf':
            return self._generate_balance_sheet_pdf(data, as_of_date)
        else:
            return self._generate_balance_sheet_excel(data, as_of_date)
    
    def _get_trial_balance_data(self, from_date: Optional[str], to_date: Optional[str]):
        with db_manager.get_session() as session:
            query = session.query(
                AccountMaster.id,
                AccountMaster.name,
                AccountMaster.code,
                AccountGroup.account_type,
                func.sum(Ledger.debit_amount).label('total_debit'),
                func.sum(Ledger.credit_amount).label('total_credit')
            ).join(AccountGroup).outerjoin(Ledger).filter(
                AccountMaster.tenant_id == self.session_manager.tenant_id
            )
            
            if from_date:
                query = query.filter(Ledger.transaction_date >= datetime.fromisoformat(from_date))
            if to_date:
                query = query.filter(Ledger.transaction_date <= datetime.fromisoformat(to_date))
            
            query = query.group_by(AccountMaster.id, AccountMaster.name, AccountMaster.code, AccountGroup.account_type)
            
            results = query.all()
            
            accounts = []
            grand_total_debit = 0
            grand_total_credit = 0
            
            for row in results:
                debit = float(row.total_debit or 0)
                credit = float(row.total_credit or 0)
                balance = debit - credit
                
                accounts.append({
                    "account_name": row.name,
                    "account_code": row.code,
                    "account_type": row.account_type,
                    "debit": debit,
                    "credit": credit,
                    "balance": balance
                })
                
                grand_total_debit += debit
                grand_total_credit += credit
            
            return {
                "accounts": accounts,
                "grand_total_debit": grand_total_debit,
                "grand_total_credit": grand_total_credit,
                "difference": grand_total_debit - grand_total_credit
            }
    
    def _get_profit_loss_data(self, from_date: Optional[str], to_date: Optional[str]):
        with db_manager.get_session() as session:
            query = session.query(
                AccountMaster.name,
                AccountGroup.account_type,
                func.sum(Ledger.debit_amount).label('total_debit'),
                func.sum(Ledger.credit_amount).label('total_credit')
            ).join(AccountGroup).join(Ledger).filter(
                AccountMaster.tenant_id == self.session_manager.tenant_id,
                AccountGroup.account_type.in_(['INCOME', 'EXPENSE'])
            )
            
            if from_date:
                query = query.filter(Ledger.transaction_date >= datetime.fromisoformat(from_date))
            if to_date:
                query = query.filter(Ledger.transaction_date <= datetime.fromisoformat(to_date))
            
            query = query.group_by(AccountMaster.name, AccountGroup.account_type)
            
            results = query.all()
            
            income_accounts = []
            expense_accounts = []
            total_income = 0
            total_expense = 0
            
            for row in results:
                debit = float(row.total_debit or 0)
                credit = float(row.total_credit or 0)
                
                if row.account_type == 'INCOME':
                    amount = credit - debit
                    income_accounts.append({"name": row.name, "amount": amount})
                    total_income += amount
                else:
                    amount = debit - credit
                    expense_accounts.append({"name": row.name, "amount": amount})
                    total_expense += amount
            
            return {
                "income": income_accounts,
                "expenses": expense_accounts,
                "total_income": total_income,
                "total_expense": total_expense,
                "net_profit": total_income - total_expense
            }
    
    def _get_balance_sheet_data(self, as_of_date: Optional[str]):
        with db_manager.get_session() as session:
            query = session.query(
                AccountMaster.name,
                AccountGroup.account_type,
                func.sum(Ledger.debit_amount).label('total_debit'),
                func.sum(Ledger.credit_amount).label('total_credit')
            ).join(AccountGroup).outerjoin(Ledger).filter(
                AccountMaster.tenant_id == self.session_manager.tenant_id,
                AccountGroup.account_type.in_(['ASSET', 'LIABILITY', 'EQUITY'])
            )
            
            if as_of_date:
                query = query.filter(Ledger.transaction_date <= datetime.fromisoformat(as_of_date))
            
            query = query.group_by(AccountMaster.name, AccountGroup.account_type)
            
            results = query.all()
            
            assets = []
            liabilities = []
            equity = []
            total_assets = 0
            total_liabilities = 0
            total_equity = 0
            
            for row in results:
                debit = float(row.total_debit or 0)
                credit = float(row.total_credit or 0)
                balance = debit - credit
                
                if row.account_type == 'ASSET':
                    assets.append({"name": row.name, "amount": balance})
                    total_assets += balance
                elif row.account_type == 'LIABILITY':
                    liabilities.append({"name": row.name, "amount": -balance})
                    total_liabilities += -balance
                else:
                    equity.append({"name": row.name, "amount": -balance})
                    total_equity += -balance
            
            return {
                "assets": assets,
                "liabilities": liabilities,
                "equity": equity,
                "total_assets": total_assets,
                "total_liabilities": total_liabilities,
                "total_equity": total_equity
            }
    
    def _generate_trial_balance_pdf(self, data, from_date, to_date) -> Tuple[bytes, str, str]:
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER)
            elements.append(Paragraph("Trial Balance", title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Date range
            date_text = f"Period: {from_date or 'Beginning'} to {to_date or 'Current'}"
            elements.append(Paragraph(date_text, styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Table data
            table_data = [['Account Code', 'Account Name', 'Debit', 'Credit', 'Balance']]
            for acc in data['accounts']:
                table_data.append([
                    acc['account_code'],
                    acc['account_name'],
                    f"₹{acc['debit']:,.2f}",
                    f"₹{acc['credit']:,.2f}",
                    f"₹{acc['balance']:,.2f}"
                ])
            
            table_data.append(['', 'Total', f"₹{data['grand_total_debit']:,.2f}", 
                             f"₹{data['grand_total_credit']:,.2f}", ''])
            
            table = Table(table_data, colWidths=[1*inch, 3*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            filename = f"trial_balance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return buffer.getvalue(), filename, "application/pdf"
        except ImportError:
            # Fallback if reportlab not installed
            return b"PDF generation requires reportlab library", "error.txt", "text/plain"
    
    def _generate_trial_balance_excel(self, data, from_date, to_date) -> Tuple[bytes, str, str]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Trial Balance"
            
            # Title
            ws['A1'] = "Trial Balance"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Period: {from_date or 'Beginning'} to {to_date or 'Current'}"
            
            # Headers
            headers = ['Account Code', 'Account Name', 'Debit', 'Credit', 'Balance']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data
            row = 5
            for acc in data['accounts']:
                ws.cell(row=row, column=1, value=acc['account_code'])
                ws.cell(row=row, column=2, value=acc['account_name'])
                ws.cell(row=row, column=3, value=acc['debit'])
                ws.cell(row=row, column=4, value=acc['credit'])
                ws.cell(row=row, column=5, value=acc['balance'])
                row += 1
            
            # Totals
            ws.cell(row=row, column=2, value="Total").font = Font(bold=True)
            ws.cell(row=row, column=3, value=data['grand_total_debit']).font = Font(bold=True)
            ws.cell(row=row, column=4, value=data['grand_total_credit']).font = Font(bold=True)
            
            # Column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            buffer = BytesIO()
            wb.save(buffer)
            filename = f"trial_balance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return buffer.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        except ImportError:
            # Fallback if openpyxl not installed
            return b"Excel generation requires openpyxl library", "error.txt", "text/plain"
    
    def _generate_profit_loss_pdf(self, data, from_date, to_date) -> Tuple[bytes, str, str]:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER)
            elements.append(Paragraph("Profit & Loss Statement", title_style))
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph(f"Period: {from_date or 'Beginning'} to {to_date or 'Current'}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Income section
            table_data = [['Income', 'Amount']]
            for item in data['income']:
                table_data.append([item['name'], f"₹{item['amount']:,.2f}"])
            table_data.append(['Total Income', f"₹{data['total_income']:,.2f}"])
            
            # Expense section
            table_data.append(['', ''])
            table_data.append(['Expenses', 'Amount'])
            for item in data['expenses']:
                table_data.append([item['name'], f"₹{item['amount']:,.2f}"])
            table_data.append(['Total Expenses', f"₹{data['total_expense']:,.2f}"])
            
            # Net profit
            table_data.append(['', ''])
            table_data.append(['Net Profit', f"₹{data['net_profit']:,.2f}"])
            
            table = Table(table_data, colWidths=[4*inch, 2*inch])
            table.setStyle(TableStyle([
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen if data['net_profit'] >= 0 else colors.lightcoral),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            filename = f"profit_loss_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return buffer.getvalue(), filename, "application/pdf"
        except ImportError:
            return b"PDF generation requires reportlab library", "error.txt", "text/plain"
    
    def _generate_profit_loss_excel(self, data, from_date, to_date) -> Tuple[bytes, str, str]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Profit & Loss"
            
            ws['A1'] = "Profit & Loss Statement"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Period: {from_date or 'Beginning'} to {to_date or 'Current'}"
            
            row = 4
            ws.cell(row=row, column=1, value="Income").font = Font(bold=True)
            row += 1
            for item in data['income']:
                ws.cell(row=row, column=1, value=item['name'])
                ws.cell(row=row, column=2, value=item['amount'])
                row += 1
            ws.cell(row=row, column=1, value="Total Income").font = Font(bold=True)
            ws.cell(row=row, column=2, value=data['total_income']).font = Font(bold=True)
            
            row += 2
            ws.cell(row=row, column=1, value="Expenses").font = Font(bold=True)
            row += 1
            for item in data['expenses']:
                ws.cell(row=row, column=1, value=item['name'])
                ws.cell(row=row, column=2, value=item['amount'])
                row += 1
            ws.cell(row=row, column=1, value="Total Expenses").font = Font(bold=True)
            ws.cell(row=row, column=2, value=data['total_expense']).font = Font(bold=True)
            
            row += 2
            ws.cell(row=row, column=1, value="Net Profit").font = Font(bold=True, size=14)
            ws.cell(row=row, column=2, value=data['net_profit']).font = Font(bold=True, size=14)
            
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            
            buffer = BytesIO()
            wb.save(buffer)
            filename = f"profit_loss_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return buffer.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        except ImportError:
            return b"Excel generation requires openpyxl library", "error.txt", "text/plain"
    
    def _generate_balance_sheet_pdf(self, data, as_of_date) -> Tuple[bytes, str, str]:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER)
            elements.append(Paragraph("Balance Sheet", title_style))
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph(f"As of: {as_of_date or 'Current Date'}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            table_data = [['Assets', 'Amount']]
            for item in data['assets']:
                table_data.append([item['name'], f"₹{item['amount']:,.2f}"])
            table_data.append(['Total Assets', f"₹{data['total_assets']:,.2f}"])
            
            table_data.append(['', ''])
            table_data.append(['Liabilities', 'Amount'])
            for item in data['liabilities']:
                table_data.append([item['name'], f"₹{item['amount']:,.2f}"])
            table_data.append(['Total Liabilities', f"₹{data['total_liabilities']:,.2f}"])
            
            table_data.append(['', ''])
            table_data.append(['Equity', 'Amount'])
            for item in data['equity']:
                table_data.append([item['name'], f"₹{item['amount']:,.2f}"])
            table_data.append(['Total Equity', f"₹{data['total_equity']:,.2f}"])
            
            table = Table(table_data, colWidths=[4*inch, 2*inch])
            table.setStyle(TableStyle([
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            filename = f"balance_sheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return buffer.getvalue(), filename, "application/pdf"
        except ImportError:
            return b"PDF generation requires reportlab library", "error.txt", "text/plain"
    
    def _generate_balance_sheet_excel(self, data, as_of_date) -> Tuple[bytes, str, str]:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Balance Sheet"
            
            ws['A1'] = "Balance Sheet"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"As of: {as_of_date or 'Current Date'}"
            
            row = 4
            ws.cell(row=row, column=1, value="Assets").font = Font(bold=True)
            row += 1
            for item in data['assets']:
                ws.cell(row=row, column=1, value=item['name'])
                ws.cell(row=row, column=2, value=item['amount'])
                row += 1
            ws.cell(row=row, column=1, value="Total Assets").font = Font(bold=True)
            ws.cell(row=row, column=2, value=data['total_assets']).font = Font(bold=True)
            
            row += 2
            ws.cell(row=row, column=1, value="Liabilities").font = Font(bold=True)
            row += 1
            for item in data['liabilities']:
                ws.cell(row=row, column=1, value=item['name'])
                ws.cell(row=row, column=2, value=item['amount'])
                row += 1
            ws.cell(row=row, column=1, value="Total Liabilities").font = Font(bold=True)
            ws.cell(row=row, column=2, value=data['total_liabilities']).font = Font(bold=True)
            
            row += 2
            ws.cell(row=row, column=1, value="Equity").font = Font(bold=True)
            row += 1
            for item in data['equity']:
                ws.cell(row=row, column=1, value=item['name'])
                ws.cell(row=row, column=2, value=item['amount'])
                row += 1
            ws.cell(row=row, column=1, value="Total Equity").font = Font(bold=True)
            ws.cell(row=row, column=2, value=data['total_equity']).font = Font(bold=True)
            
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            
            buffer = BytesIO()
            wb.save(buffer)
            filename = f"balance_sheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return buffer.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        except ImportError:
            return b"Excel generation requires openpyxl library", "error.txt", "text/plain"
